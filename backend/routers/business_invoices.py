"""
İşletme Faturaları (Alınan Faturalar) API
Restoran Raporu Excel'inden işletme fatura tutarlarını import eder
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from datetime import datetime, timezone, timedelta
from typing import Optional
import uuid
import base64
import openpyxl
from io import BytesIO

from utils.database import db
from utils.helpers import get_turkey_now, ensure_turkey_timezone, TURKEY_TZ
from services.r2_storage import upload_file_to_r2
from utils.jwt_utils import require_admin

router = APIRouter(prefix="/api/business-invoices", tags=["İşletme Faturaları"], dependencies=[Depends(require_admin)])


# =====================================================
# KESİLEN FATURALAR (ISSUED INVOICES) API - Must be before dynamic routes
# =====================================================

# --- Get all issued invoice records for a company (independent of month) ---
@router.get("/get-all-issued/{company_id}")
async def get_all_issued_invoices(company_id: str):
    """Get all issued invoice records for a company (not filtered by month)"""
    records = await db.issued_invoices.find(
        {"company_id": company_id},
        {"_id": 0}
    ).to_list(500)
    
    return records


# --- Mark invoice as issued for a business (independent of month) ---
@router.post("/mark-issued/{company_id}/{business_id}")
async def mark_invoice_issued(company_id: str, business_id: str):
    """Mark invoice as issued - saves Monday of current week as date"""
    # Get business name
    business = await db.businesses.find_one({"id": business_id}, {"_id": 0, "name": 1})
    if not business:
        raise HTTPException(status_code=404, detail="İşletme bulunamadı")
    
    # Calculate Monday of current week
    today = datetime.now(TURKEY_TZ)
    days_since_monday = today.weekday()  # Monday = 0, Sunday = 6
    monday = today.replace(hour=0, minute=0, second=0, microsecond=0)
    if days_since_monday > 0:
        monday = monday - timedelta(days=days_since_monday)
    
    monday_str = monday.strftime("%Y-%m-%d")
    
    # Check if business already has issued record
    existing = await db.issued_invoices.find_one({
        "company_id": company_id,
        "business_id": business_id
    })
    
    if existing:
        await db.issued_invoices.update_one(
            {"id": existing["id"]},
            {"$set": {
                "issued_until_date": monday_str,
                "updated_at": get_turkey_now()
            }}
        )
    else:
        record = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "business_id": business_id,
            "business_name": business["name"],
            "issued_until_date": monday_str,
            "created_at": get_turkey_now()
        }
        await db.issued_invoices.insert_one(record)
    
    return {"message": "Fatura kesildi olarak işaretlendi", "issued_until_date": monday_str}


# --- Clear issued status for a business ---
@router.delete("/clear-issued/{company_id}/{business_id}")
async def clear_issued_invoice(company_id: str, business_id: str):
    """Clear the issued status for a business"""
    result = await db.issued_invoices.delete_one({
        "company_id": company_id,
        "business_id": business_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    
    return {"message": "İşaret kaldırıldı"}


# =====================================================
# ALINAN FATURALAR (RECEIVED INVOICES) API
# =====================================================

# --- Get all business invoice records for a month ---
@router.get("/{company_id}/{year}/{month}")
async def get_business_invoices(company_id: str, year: int, month: int):
    """Get all business invoice records for a specific month"""
    records = await db.business_invoices.find(
        {"company_id": company_id, "year": year, "month": month},
        {"_id": 0}
    ).to_list(500)
    
    return records


# --- Download all invoices for a month (bulk) - Must be before /{business_id} ---
@router.get("/{company_id}/{year}/{month}/download-all")
async def download_all_invoices(company_id: str, year: int, month: int):
    """Get all invoice files for the month as a list"""
    records = await db.business_invoices.find(
        {
            "company_id": company_id,
            "year": year,
            "month": month,
            "invoice_uploaded": True
        },
        {"_id": 0}
    ).to_list(500)
    
    all_invoices = []
    
    for record in records:
        business_name = record.get("business_name", "Bilinmeyen")
        invoices = record.get("invoices", [])
        
        # Handle old format
        if not invoices and record.get("invoice_file"):
            invoices = [{
                "invoice_id": "legacy",
                "file_data": record["invoice_file"],
                "filename": record.get("invoice_filename", "fatura.pdf"),
                "extension": record.get("invoice_extension", "pdf")
            }]
        
        for inv in invoices:
            all_invoices.append({
                "business_name": business_name,
                "business_id": record.get("business_id"),
                "invoice_id": inv.get("invoice_id"),
                "file_data": inv.get("file_data"),
                "filename": inv.get("filename"),
                "extension": inv.get("extension")
            })
    
    return {"invoices": all_invoices, "count": len(all_invoices)}


# --- Get single business invoice record ---
@router.get("/{company_id}/{year}/{month}/{business_id}")
async def get_business_invoice(company_id: str, year: int, month: int, business_id: str):
    """Get a specific business invoice record"""
    record = await db.business_invoices.find_one(
        {
            "company_id": company_id,
            "year": year,
            "month": month,
            "business_id": business_id
        },
        {"_id": 0}
    )
    
    return record


# --- Upload Excel and import invoice amounts ---
@router.post("/{company_id}/import-excel")
async def import_excel(
    company_id: str,
    year: int = Form(...),
    month: int = Form(...),
    file: UploadFile = File(...)
):
    """
    Import business invoice amounts from Restoran Raporu.xlsx
    Reads 'Restoran Raporu' (or 'Restoran Adı') and 'Banka/Kredi Kartı' columns
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyası (.xlsx, .xls) yüklenebilir")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        
        # Find header row - look for "Restoran Raporu" or "Restoran Adı" in first 10 rows
        header_row = None
        restoran_col = None
        banka_col = None
        
        for row_idx in range(1, 11):
            for col_idx in range(1, 20):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.strip().lower()
                    # Match "Restoran Raporu" or "Restoran Adı"
                    if "restoran" in cell_lower and ("rapor" in cell_lower or "ad" in cell_lower):
                        header_row = row_idx
                        restoran_col = col_idx
                    # Match "Banka/Kredi Kartı"
                    elif "banka" in cell_lower and "kredi" in cell_lower:
                        banka_col = col_idx
            
            if header_row and restoran_col and banka_col:
                break
        
        if not header_row or not restoran_col or not banka_col:
            raise HTTPException(
                status_code=400, 
                detail="Excel dosyasında 'Restoran Raporu' ve 'Banka/Kredi Kartı' sütunları bulunamadı"
            )
        
        # Get all businesses for matching
        businesses = await db.businesses.find(
            {"company_id": company_id, "is_archived": {"$ne": True}},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(500)
        
        business_map = {b["name"].strip().lower(): b for b in businesses}
        
        # Parse data rows
        imported_count = 0
        not_found = []
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            restoran_name = ws.cell(row=row_idx, column=restoran_col).value
            banka_value = ws.cell(row=row_idx, column=banka_col).value
            
            if not restoran_name:
                continue
            
            restoran_name_clean = str(restoran_name).strip()
            restoran_name_lower = restoran_name_clean.lower()
            
            # Try to match business
            matched_business = business_map.get(restoran_name_lower)
            
            if not matched_business:
                # Try partial match
                for bname, bdata in business_map.items():
                    if restoran_name_lower in bname or bname in restoran_name_lower:
                        matched_business = bdata
                        break
            
            if not matched_business:
                not_found.append(restoran_name_clean)
                continue
            
            # Parse amount - handle Turkish number format (₺1.234,56)
            amount = 0.0
            if banka_value is not None:
                try:
                    if isinstance(banka_value, (int, float)):
                        amount = float(banka_value)
                    else:
                        # Clean string - Turkish format: ₺1.234,56 -> 1234.56
                        amount_str = str(banka_value).strip()
                        # Remove currency symbols and spaces
                        amount_str = amount_str.replace('₺', '').replace('TL', '').replace(' ', '')
                        # Handle Turkish number format: remove thousand separator (.) then replace decimal (,) with (.)
                        if ',' in amount_str:
                            # Turkish format: 1.234,56 -> remove dots, replace comma with dot
                            amount_str = amount_str.replace('.', '').replace(',', '.')
                        # Handle negative values like -₺55.021,09
                        amount = float(amount_str)
                except (ValueError, TypeError):
                    amount = 0.0
            
            # Update or create record
            existing = await db.business_invoices.find_one({
                "company_id": company_id,
                "year": year,
                "month": month,
                "business_id": matched_business["id"]
            })
            
            if existing:
                await db.business_invoices.update_one(
                    {"id": existing["id"]},
                    {"$set": {
                        "required_amount": amount,
                        "business_name": matched_business["name"],
                        "updated_at": get_turkey_now()
                    }}
                )
            else:
                record = {
                    "id": str(uuid.uuid4()),
                    "company_id": company_id,
                    "year": year,
                    "month": month,
                    "business_id": matched_business["id"],
                    "business_name": matched_business["name"],
                    "required_amount": amount,
                    "invoice_uploaded": False,
                    "invoice_file": None,
                    "invoice_filename": None,
                    "uploaded_at": None,
                    "created_at": get_turkey_now()
                }
                await db.business_invoices.insert_one(record)
            
            imported_count += 1
        
        return {
            "message": f"{imported_count} işletme için fatura tutarı aktarıldı",
            "imported_count": imported_count,
            "not_found": not_found[:10],  # First 10 not found
            "not_found_count": len(not_found)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel işleme hatası: {str(e)}")


# --- Upload invoice PDF for a business (supports multiple invoices) ---
@router.post("/{company_id}/{year}/{month}/{business_id}/upload")
async def upload_invoice(
    company_id: str,
    year: int,
    month: int,
    business_id: str,
    file: UploadFile = File(...)
):
    """Upload the received invoice PDF from a business - supports multiple invoices"""
    if not file.filename.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png')):
        raise HTTPException(status_code=400, detail="Sadece PDF veya resim dosyası yüklenebilir")
    
    # Read file content
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="Dosya boyutu 10MB'dan büyük olamaz")
    
    # Upload to R2
    file_extension = file.filename.split('.')[-1].lower()
    r2_key = f"business-invoices/{company_id}/{year}/{month}/{business_id}/{str(uuid.uuid4())}.{file_extension}"
    content_type = file.content_type or "application/pdf"
    
    upload_result = await upload_file_to_r2(content, r2_key, content_type)
    if not upload_result['success']:
        raise HTTPException(status_code=503, detail="Dosya depolama servisi (Cloudflare R2) yapılandırılmamış. Sistem ayarlarından R2 bağlantısını yapın.")
    
    # Create invoice object
    invoice_obj = {
        "invoice_id": str(uuid.uuid4()),
        "r2_key": r2_key,
        "storage_type": "r2",
        "filename": file.filename,
        "extension": file_extension,
        "uploaded_at": get_turkey_now()
    }
    
    # Find existing record
    existing = await db.business_invoices.find_one({
        "company_id": company_id,
        "year": year,
        "month": month,
        "business_id": business_id
    })
    
    if existing:
        # Get existing invoices list or migrate from old format
        invoices = existing.get("invoices", [])
        
        # Migrate old single invoice to list format if needed
        if not invoices and existing.get("invoice_file"):
            old_invoice = {
                "invoice_id": str(uuid.uuid4()),
                "file_data": existing["invoice_file"],
                "filename": existing.get("invoice_filename", "fatura.pdf"),
                "extension": existing.get("invoice_extension", "pdf"),
                "uploaded_at": existing.get("uploaded_at", get_turkey_now())
            }
            invoices.append(old_invoice)
        
        # Add new invoice
        invoices.append(invoice_obj)
        
        await db.business_invoices.update_one(
            {"id": existing["id"]},
            {"$set": {
                "invoice_uploaded": True,
                "invoices": invoices,
                # Keep old fields for backward compatibility but clear them
                "invoice_file": None,
                "invoice_filename": None,
                "invoice_extension": None,
                "uploaded_at": get_turkey_now()
            }}
        )
    else:
        # Get business name
        business = await db.businesses.find_one({"id": business_id}, {"_id": 0, "name": 1})
        business_name = business["name"] if business else "Bilinmeyen"
        
        record = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "year": year,
            "month": month,
            "business_id": business_id,
            "business_name": business_name,
            "required_amount": 0,
            "invoice_uploaded": True,
            "invoices": [invoice_obj],
            "uploaded_at": get_turkey_now(),
            "created_at": get_turkey_now()
        }
        await db.business_invoices.insert_one(record)
    
    return {"message": "Fatura yüklendi"}


# --- Download invoice file ---
@router.get("/{company_id}/{year}/{month}/{business_id}/download/{invoice_id}")
async def download_invoice(company_id: str, year: int, month: int, business_id: str, invoice_id: str):
    """Download a specific invoice file"""
    record = await db.business_invoices.find_one({
        "company_id": company_id,
        "year": year,
        "month": month,
        "business_id": business_id
    })
    
    if not record:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    
    invoices = record.get("invoices", [])
    
    # Handle old single invoice format
    if not invoices and record.get("invoice_file"):
        return {
            "file_data": record["invoice_file"],
            "filename": record.get("invoice_filename", "fatura.pdf"),
            "extension": record.get("invoice_extension", "pdf")
        }
    
    # Find specific invoice
    for inv in invoices:
        if inv.get("invoice_id") == invoice_id:
            return {
                "file_data": inv["file_data"],
                "filename": inv.get("filename", "fatura.pdf"),
                "extension": inv.get("extension", "pdf")
            }
    
    raise HTTPException(status_code=404, detail="Fatura bulunamadı")


# --- Delete a specific invoice ---
@router.delete("/{company_id}/{year}/{month}/{business_id}/invoice/{invoice_id}")
async def delete_invoice(company_id: str, year: int, month: int, business_id: str, invoice_id: str):
    """Delete a specific invoice file"""
    record = await db.business_invoices.find_one({
        "company_id": company_id,
        "year": year,
        "month": month,
        "business_id": business_id
    })
    
    if not record:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    
    invoices = record.get("invoices", [])
    
    # Filter out the invoice to delete
    updated_invoices = [inv for inv in invoices if inv.get("invoice_id") != invoice_id]
    
    if len(updated_invoices) == len(invoices):
        raise HTTPException(status_code=404, detail="Fatura bulunamadı")
    
    await db.business_invoices.update_one(
        {"id": record["id"]},
        {"$set": {
            "invoices": updated_invoices,
            "invoice_uploaded": len(updated_invoices) > 0
        }}
    )
    
    return {"message": "Fatura silindi"}


# --- Get company invoice details from system settings ---
@router.get("/company-details/{company_id}")
async def get_company_invoice_details(company_id: str):
    """Get company invoice details for WhatsApp message"""
    company = await db.companies.find_one(
        {"id": company_id},
        {"_id": 0, "name": 1, "tckn_vkn": 1, "address": 1, "tax_office": 1, "email": 1}
    )
    
    if not company:
        raise HTTPException(status_code=404, detail="Şirket bulunamadı")
    
    return company


# --- Set manual amount for a business ---
@router.post("/{company_id}/{year}/{month}/{business_id}/set-amount")
async def set_amount(
    company_id: str,
    year: int,
    month: int,
    business_id: str,
    data: dict
):
    """Manually set invoice amount for a business"""
    amount = data.get("amount", 0)
    
    # Get business name
    business = await db.businesses.find_one({"id": business_id}, {"_id": 0, "name": 1})
    if not business:
        raise HTTPException(status_code=404, detail="İşletme bulunamadı")
    
    existing = await db.business_invoices.find_one({
        "company_id": company_id,
        "year": year,
        "month": month,
        "business_id": business_id
    })
    
    if existing:
        await db.business_invoices.update_one(
            {"id": existing["id"]},
            {"$set": {
                "required_amount": amount,
                "updated_at": get_turkey_now()
            }}
        )
    else:
        record = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "year": year,
            "month": month,
            "business_id": business_id,
            "business_name": business["name"],
            "required_amount": amount,
            "invoice_uploaded": False,
            "invoice_file": None,
            "invoice_filename": None,
            "uploaded_at": None,
            "created_at": get_turkey_now()
        }
        await db.business_invoices.insert_one(record)
    
    return {"message": "Tutar kaydedildi"}
