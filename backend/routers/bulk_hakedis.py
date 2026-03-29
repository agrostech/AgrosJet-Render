from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import io
import re

from utils.database import db
from utils.helpers import get_turkey_now, ensure_turkey_timezone, TURKEY_TZ
from utils.jwt_utils import require_admin

router = APIRouter(prefix="/api/bulk-hakedis", tags=["Toplu Hakediş"], dependencies=[Depends(require_admin)])


class BulkHakedisItem(BaseModel):
    courier_id: str
    courier_name: str
    hakedis_amount: float
    packet_count: int
    bonus_amount: float


class BulkHakedisCreate(BaseModel):
    items: List[BulkHakedisItem]
    admin_id: str
    admin_name: str
    custom_date: Optional[str] = None
    add_jetpuan: Optional[bool] = True  # JetPuan eklensin mi?


def parse_turkish_number(value) -> float:
    """Parse number, handling both Turkish format (1.234,56) and standard floats"""
    if value is None:
        return 0
    
    # If already a number, return directly
    if isinstance(value, (int, float)):
        return float(value)
    
    # Convert to string for parsing
    value_str = str(value).strip()
    if not value_str:
        return 0
    
    # Remove currency symbol and whitespace
    cleaned = re.sub(r'[₺TL\s]', '', value_str)
    
    # Check if it's Turkish format (has comma as decimal separator)
    if ',' in cleaned and '.' in cleaned:
        # Turkish format: 1.234,56 -> 1234.56
        cleaned = cleaned.replace('.', '').replace(',', '.')
    elif ',' in cleaned:
        # Only comma (could be Turkish decimal): 1234,56 -> 1234.56
        cleaned = cleaned.replace(',', '.')
    # If only dots, it's standard format, keep as is
    
    try:
        return float(cleaned)
    except:
        return 0


def parse_packet_count(value) -> int:
    """Parse packet count, handling empty values"""
    if not value:
        return 0
    try:
        return int(float(str(value).replace(',', '.')))
    except:
        return 0


@router.post("/parse-excel/{company_id}")
async def parse_excel(company_id: str, file: UploadFile = File(...)):
    """Parse Excel file and match courier names"""
    import openpyxl
    
    # Read Excel file
    contents = await file.read()
    
    # Boyut kontrolü - Excel max 5MB
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Excel dosyası 5MB'ı geçemez")
    
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # Get all couriers for this company
    company_couriers = await db.company_couriers.find(
        {"company_id": company_id},
        {"_id": 0, "courier_id": 1}
    ).to_list(1000)
    courier_ids = [cc["courier_id"] for cc in company_couriers]
    
    couriers = await db.couriers.find(
        {"id": {"$in": courier_ids}, "is_archived": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(1000)
    
    # Create name lookup (lowercase for matching)
    courier_lookup = {}
    for c in couriers:
        name_lower = c["name"].lower().strip()
        courier_lookup[name_lower] = c
    
    # Get bonus rules
    bonus_rules = await db.bonus_settings.find(
        {"company_id": company_id},
        {"_id": 0}
    ).sort("min_packets", -1).to_list(100)
    
    def calculate_bonus(packet_count: int) -> float:
        for rule in bonus_rules:
            if packet_count >= rule["min_packets"]:
                return rule["amount"]
        return 0
    
    # Parse Excel rows
    matched = []
    unmatched = []
    
    # Find header row and column indices
    header_row = None
    name_col = None
    packet_col = None
    hakedis_col = None
    
    # Search for header row - must have at least 2 recognized columns to be valid header
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        temp_name_col = None
        temp_packet_col = None
        temp_hakedis_col = None
        
        for col_idx, cell in enumerate(row):
            if cell:
                cell_str = str(cell).lower().strip()
                # Check for name column - but must be exactly "kurye" not "kurye raporu"
                if cell_str == 'kurye' or cell_str == 'isim' or cell_str == 'ad' or cell_str == 'kurye adı':
                    temp_name_col = col_idx
                elif 'paket' in cell_str:
                    temp_packet_col = col_idx
                elif 'kazanç' in cell_str or 'hakediş' in cell_str or 'hakedis' in cell_str or cell_str == 'total':
                    temp_hakedis_col = col_idx
        
        # Valid header row must have name column AND at least one other column
        if temp_name_col is not None and (temp_packet_col is not None or temp_hakedis_col is not None):
            name_col = temp_name_col
            packet_col = temp_packet_col
            hakedis_col = temp_hakedis_col
            header_row = row_idx
            break
    
    if not header_row or name_col is None:
        raise HTTPException(status_code=400, detail="Excel formatı tanınamadı. 'Kurye' veya 'İsim' sütunu bulunamadı.")
    
    # Parse data rows
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[name_col] if name_col is not None and len(row) > name_col else None
        if not name or str(name).strip() == '':
            continue
        
        name_str = str(name).strip()
        name_lower = name_str.lower()
        
        packet_count = 0
        if packet_col is not None and len(row) > packet_col:
            packet_count = parse_packet_count(row[packet_col])
        
        hakedis_amount = 0
        if hakedis_col is not None and len(row) > hakedis_col:
            hakedis_amount = parse_turkish_number(row[hakedis_col])
        
        bonus_amount = calculate_bonus(packet_count)
        
        # Try to match courier
        matched_courier = courier_lookup.get(name_lower)
        
        if matched_courier:
            matched.append({
                "courier_id": matched_courier["id"],
                "courier_name": matched_courier["name"],
                "excel_name": name_str,
                "packet_count": packet_count,
                "hakedis_amount": hakedis_amount,
                "bonus_amount": bonus_amount,
                "total_amount": hakedis_amount + bonus_amount
            })
        else:
            unmatched.append({
                "excel_name": name_str,
                "packet_count": packet_count,
                "hakedis_amount": hakedis_amount,
                "bonus_amount": bonus_amount
            })
    
    return {
        "matched": matched,
        "unmatched": unmatched,
        "total_matched": len(matched),
        "total_unmatched": len(unmatched),
        "bonus_rules": bonus_rules
    }


@router.post("/apply/{company_id}")
async def apply_bulk_hakedis(company_id: str, data: BulkHakedisCreate):
    """Apply bulk hakediş to matched couriers"""
    from routers.jetpuan import calculate_and_credit_points
    from routers.notifications import create_notification
    from routers.accounting import create_activity_log, get_admin_role
    
    # Admin rolünü al
    admin_role = await get_admin_role(data.admin_id)
    
    # Parse custom date
    if data.custom_date:
        try:
            tx_date = datetime.fromisoformat(data.custom_date.replace('Z', '+00:00'))
            if tx_date.tzinfo is None:
                tx_date = tx_date.replace(tzinfo=timezone.utc)
            created_at = tx_date.isoformat()
        except:
            created_at = get_turkey_now()
    else:
        created_at = get_turkey_now()
    
    results = []
    
    for item in data.items:
        total_amount = item.hakedis_amount + item.bonus_amount
        
        # Build description - only include bonus if earned
        desc_parts = []
        if item.packet_count > 0:
            desc_parts.append(f"{item.packet_count} Paket")
        if item.hakedis_amount > 0:
            desc_parts.append(f"{item.hakedis_amount:.2f} TL Hakediş")
        if item.bonus_amount > 0:
            desc_parts.append(f"{item.bonus_amount:.2f} TL Bonus")
        
        description = ", ".join(desc_parts) if desc_parts else "Toplu Hakediş"
        
        # Create transaction
        transaction = {
            "id": str(uuid.uuid4()),
            "entity_type": "courier",
            "entity_id": item.courier_id,
            "company_id": company_id,
            "type": "payment_in",  # Alınan
            "amount": total_amount,
            "description": description,
            "is_hakedis": True,
            "admin_id": data.admin_id,
            "admin_name": data.admin_name,
            "created_at": created_at
        }
        await db.transactions.insert_one(transaction)
        
        # Credit JetPuan for hakediş (only if add_jetpuan is True)
        if data.add_jetpuan:
            try:
                await calculate_and_credit_points(item.courier_id, total_amount)
            except Exception as e:
                print(f"JetPuan credit failed: {e}")
        
        # Create activity log
        try:
            await create_activity_log({
                "company_id": company_id,
                "admin_id": data.admin_id,
                "admin_name": data.admin_name,
                "action": "bulk_hakedis",
                "entity_type": "courier",
                "entity_id": item.courier_id,
                "entity_name": item.courier_name,
                "details": {
                    "transaction_id": transaction["id"],
                    "type": "payment_in",
                    "amount": total_amount,
                    "hakedis_amount": item.hakedis_amount,
                    "bonus_amount": item.bonus_amount,
                    "packet_count": item.packet_count,
                    "description": description
                }
            })
        except Exception as e:
            print(f"Activity log failed: {e}")
        
        results.append({
            "courier_id": item.courier_id,
            "courier_name": item.courier_name,
            "amount": total_amount,
            "transaction_id": transaction["id"]
        })
    
    # Send notification (superadmin hariç)
    if admin_role != "superadmin":
        try:
            await create_notification(
                company_id=company_id,
                notification_type="bulk_hakedis",
                title="Toplu Hakediş Eklendi",
                message=f"{len(results)} kuryeye toplu hakediş eklendi.",
                entity_type="bulk_hakedis",
                entity_id=str(uuid.uuid4())
            )
        except Exception as e:
            print(f"Notification failed: {e}")
    
    return {
        "message": f"{len(results)} kuryeye hakediş eklendi",
        "results": results,
        "total_amount": sum(r["amount"] for r in results)
    }
